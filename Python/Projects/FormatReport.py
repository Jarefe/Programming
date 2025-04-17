from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Side, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, Rule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import os, sys

def is_sheet_empty(sheet): # TODO optimize this function
    for row in sheet.iter_rows(min_row=2): # skip header
        for cell in row:
            if cell.value is not None:
                return False # if sheet has data beyond header
    return True # only header exists

# FORMATTING RULES
# scrap
RED_FILL = PatternFill(bgColor='FFC7CE')
SCRAP = Rule(type="expression", dxf=DifferentialStyle(fill=RED_FILL))

# ram
RAM_RULE = {
    "DDR3": "D9E1F2",  # light blue
    "DDR4": "C6EFCE",  # light green
    "DDR5": "06402B",  # dark green
}

# style for header
ORANGE_FILL = PatternFill(
    start_color='FFA500', 
    end_color='FFA500', 
    fill_type='solid')

# thin line border
BORDER = Border(
    left=Side(style="thin"), 
    right=Side(style="thin"), 
    top=Side(style="thin"), 
    bottom=Side(style="thin"))

# left alignment and vertically centered
ALIGNMENT = Alignment(
    horizontal='left',
    vertical='center'
)

# these columns will be gray if the cells are empty
GRAY_CATEGORIES = [
    "Processor",
    "Type of RAM",
    "RAM Capacity",
    "Battery",
    "Touchscreen",
    "GPU",
    "Drive Caddy",
    "Drive Capacity",
    "Drive Interface",
    "Wi-Fi",
    "Form Factor",
    "Rack Ears",
    "# of PSUs",
    "PSU Part#",
    "# of Stack Ports",
    "Stacking Port",
    "Module",
    "# of Bays",
    "HDD Form Factor",
    "# of Processors",
    "Processor(s)",
    "RAM Breakdown",
    "RAM MFGR (Apple Only)",
    "RAM OEM Part# (Apple Only)",
    "Power Supply",
    "RAID Card",
    "Management Interface",
    "Card Slot 1",
    "Card Slot 2",
    "Card Slot 3"
]

def create_table(sheet):
    # obtain farthest cell
    max_row = sheet.max_row
    max_column = sheet.max_column

    # define range of cells including header
    table_range = f'A1:{get_column_letter(max_column)}{max_row}'
    
    # create table with data range
    table = Table(displayName=f"Table_{sheet_name.replace(' ','')}", ref=table_range)
    sheet.add_table(table)

def format_header(sheet):
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = ORANGE_FILL

def format_borders(sheet):
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = BORDER
            cell.alignment = ALIGNMENT

def autofit(sheet):
    for col in sheet.columns:
        max_length = 0
        column = col[0].column
        column_letter = get_column_letter(column)

        # get longest cell width in column
        for cell in col:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        # fit column widths based on max length
        adjusted_width = max_length + 1
        sheet.column_dimensions[column_letter].width = adjusted_width

def format_column(sheet):
    print()


# TODO handle file inputs and outputs
script_dir = os.path.dirname(os.path.abspath(__file__))
excel_path = os.path.join(script_dir, "Test Output.xlsx")

# load existing workbook
# TODO change to be dynamic
try:
    original_path = 'C:/Users/test/Downloads/Report 1.xlsx' 
    original_wb = load_workbook(original_path)
except:
    print("Error loading file, stopping program")
    sys.exit(1)

# create new workbook
wb = Workbook()

# remove default sheet from wb
wb.remove(wb.active)

# copy all non empty sheets 
for sheet_name in original_wb.sheetnames:
    original_sheet = original_wb[sheet_name]

    # skip empty sheets
    if is_sheet_empty(original_sheet):
        print(f'{sheet_name} is empty; skipping')
        continue

    # create sheet in new workbook
    new_sheet = wb.create_sheet(title=sheet_name)

    # copy data from old sheet to new sheet columnwise
    for col in original_sheet.iter_cols():
        for cell in col:
            new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)



# go through each sheet in the workbook
for sheet_name in wb.sheetnames:
    
    # get current sheet
    current_sheet = wb[sheet_name]

    # placeholder
    max_row = current_sheet.max_row
    max_col = current_sheet.max_column
    
    # create table 
    create_table(current_sheet)

    # apply orange fill to header row
    format_header(current_sheet)

    # apply borders to each cell
    format_borders(current_sheet)

    # "autofit" columns
    autofit(current_sheet)
         
    # apply conditional formatting
    match sheet_name:
        case 'Dash Inventory': # manually looping because conditional formatting gets overwritten
            for row in current_sheet.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=1):
                for cell in row:
                    if cell.value == "SCRP":  
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        case "Desktops":
            SCRAP.formula = ['$C2="SCRP"']
            current_sheet.conditional_formatting.add(f"C2:C{max_row}", SCRAP)
            for mem_type, color in RAM_RULE.items():
                rule = Rule(
                    type='expression',
                    formula=[f'$K2="{mem_type}"'],
                    dxf=DifferentialStyle(fill=PatternFill(bgColor=color))
                )
                current_sheet.conditional_formatting.add(f"K2:K{max_row}", rule)

        case "Laptops":
            SCRAP.formula = ['$C2="SCRP"']
            current_sheet.conditional_formatting.add(f"C2:C{max_row}", SCRAP)
            for mem_type, color in RAM_RULE.items():
                rule = Rule(
                    type='expression',
                    formula=[f'$K2="{mem_type}"'],
                    dxf=DifferentialStyle(fill=PatternFill(bgColor=color))
                )
                current_sheet.conditional_formatting.add(f"K2:K{max_row}", rule)

        case "Networking":
            SCRAP.formula = ['$C2="SCRP"']
            current_sheet.conditional_formatting.add(f"C2:C{max_row}", SCRAP)

        case "Servers":
            SCRAP.formula = ['$C2="SCRP"']
            current_sheet.conditional_formatting.add(f"C2:C{max_row}", SCRAP)
            for mem_type, color in RAM_RULE.items():
                rule = Rule(
                    type='expression',
                    formula=[f'$N2="{mem_type}"'],
                    dxf=DifferentialStyle(fill=PatternFill(bgColor=color))
                )
                current_sheet.conditional_formatting.add(f"N2:N{max_row}", rule)

    # TODO add manual looping to format empty cells in middle of each table

try:    
    wb.save(excel_path)
except:
    print("Error saving file")