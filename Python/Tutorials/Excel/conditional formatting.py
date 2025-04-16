from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border
from openpyxl.styles.differential import DifferentialStyle
import os

script_dir = os.path.dirname(os.path.abspath(__file__))
excel_path = os.path.join(script_dir, "conditional.xlsx")

dxf = DifferentialStyle(font=Font(bold=True), fill=PatternFill(start_color='EE1111', end_color='EE1111'))
rule = Rule(type='cellIs', dxf=dxf, formula=["10"])

from openpyxl.formatting.rule import ColorScaleRule
rule = ColorScaleRule(start_type='percentile', start_value=10, start_color='FFAA0000',
                      mid_type='percentile', mid_value=50, mid_color='FF0000AA',
                      end_type='percentile', end_value=90, end_color='FF00AA00')

from openpyxl.formatting.rule import DataBarRule
rule = DataBarRule(start_type='percentile', start_value=10, end_type='percentile', end_value='90',
                   color="FF638EC6", showValue="None", minLength=None, maxLength=None)


# standard conditional formats
# average, percent, unique/duplicate, value, rank
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
wb = Workbook()
ws = wb.active

# Create fill
redFill = PatternFill(start_color='EE1111',
               end_color='EE1111',
               fill_type='solid')

# Add a two-color scale
# Takes colors in excel 'RRGGBB' style.
ws.conditional_formatting.add('A1:A10',
            ColorScaleRule(start_type='min', start_color='AA0000',
                          end_type='max', end_color='00AA00')
                          )

# Add a three-color scale
ws.conditional_formatting.add('B1:B10',
               ColorScaleRule(start_type='percentile', start_value=10, start_color='AA0000',
                           mid_type='percentile', mid_value=50, mid_color='0000AA',
                           end_type='percentile', end_value=90, end_color='00AA00')
                             )

# Add a conditional formatting based on a cell comparison
# addCellIs(range_string, operator, formula, stopIfTrue, wb, font, border, fill)
# Format if cell is less than 'formula'
ws.conditional_formatting.add('C2:C10',
            CellIsRule(operator='lessThan', formula=['C$1'], stopIfTrue=True, fill=redFill))

# Format if cell is between 'formula'
ws.conditional_formatting.add('D2:D10',
            CellIsRule(operator='between', formula=['1','5'], stopIfTrue=True, fill=redFill))

# Format using a formula
ws.conditional_formatting.add('E1:E10',
            FormulaRule(formula=['ISBLANK(E1)'], stopIfTrue=True, fill=redFill))

# Aside from the 2-color and 3-color scales, format rules take fonts, borders and fills for styling:
myFont = Font()
myBorder = Border()
ws.conditional_formatting.add('E1:E10',
            FormulaRule(formula=['E1=0'], font=myFont, border=myBorder, fill=redFill))

# Highlight cells that contain particular text by using a special formula
red_text = Font(color="9C0006")
red_fill = PatternFill(bgColor="FFC7CE")
dxf = DifferentialStyle(font=red_text, fill=red_fill)
rule = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
rule.formula = ['NOT(ISERROR(SEARCH("highlight",A1)))']
ws.conditional_formatting.add('A1:F40', rule)
wb.save(excel_path)


# formatting entire rows
ws.append(['Software', 'Developer', 'Version'])
ws.append(['Excel', 'Microsoft', '2016'])
ws.append(['openpyxl', 'Open source', '2.6'])
ws.append(['OpenOffice', 'Apache', '4.1.4'])
ws.append(['Word', 'Microsoft', '2010'])

red_fill = PatternFill(bgColor="FFC7CE")
dxf = DifferentialStyle(fill=red_fill)
r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
r.formula = ['$A2="Microsoft"']
ws.conditional_formatting.add("A1:C10", r)