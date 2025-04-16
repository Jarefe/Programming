from openpyxl import Workbook

# initialize workbook (only in memory, will not be saved anywhere until
# using wb.save())
wb = Workbook()

# get first worksheet; set to 0 by default 
ws = wb.active

# to create new worksheets
ws1 = wb.create_sheet("Sheet 1") # insert at end (default)
ws2 = wb.create_sheet("Sheet 2", 0) # insert at first position
ws3 = wb.create_sheet("Sheet 3", -1) # insert at penultimate position

# title worksheets
ws.title = "First worksheet"

# can obtain worksheet as key of workbook
print(wb["First worksheet"])

# can get names of all worksheets in book
print(wb.sheetnames)

# can loop through sheets
for sheet in wb:
    print(sheet.title)


# accessing and modifying data
# note: cells are created in memory when accessed. this means that even looping
# through empty cells will create those cells in memory unnecessarily
ws['A1'] = 'first value'
print(ws['A1'].value)

# can also use worksheet.cell()
d = ws.cell(row=4, column=2, value=10)
print(d.value)


# can access range of cells with slicing
cell_range = ws['A1':'C2']
col_range = ws['C:D']

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)

print()
for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
    for cell in col:
        print(cell)

# to iterate through all rows/cols of a file, can use worksheet.rows or .cols
print(tuple(ws.rows))
print()
print(tuple(ws.columns))

# for values only use .values
for row in ws.values:
    for value in row:
        print(value)

# can apply to iter_rows and iter_cols
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
    print(row)

# to save workbook
wb.save('Tutorial.xlsx')
# will overwrite existing files without warning

# saving as stream e.g. when using web application
# from tempfile import NamedTemporaryFile
# from openpyxl import Workbook
# wb = Workbook()
# with NamedTemporaryFile() as tmp:
#     wb.save(tmp.name)
#     tmp.seek(0)
#     stream = tmp.read()

# to load existing workbook
from openpyxl import load_workbook
try: 
    testwb = load_workbook(filename= 'empty_book.xlsx')
except:
    print("failed to load workbook")
