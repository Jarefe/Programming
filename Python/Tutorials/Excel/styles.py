from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import Workbook
import os

wb = Workbook()
ws = wb.active
script_dir = os.path.dirname(os.path.abspath(__file__))
excel_path = os.path.join(script_dir, "Styles.xlsx")

# Cell styles cannot be changed once assigned
a1 = ws['A1']
d4 = ws['D4']
a1.value = 'A1 styled'
d4.value = 'D4 styled'
ft = Font(color="FF0000")
a1.font = ft
d4.font = ft

# to change color of font, it must be reassigned
a1.font = Font(color="000000", italic=True) # only affects a1

# styles can be copied
from copy import copy

ft1 = Font(name='Arial', size=14)
ft2 = copy(ft1)
ft2.name = "Name2" # even with name changed, size is still 14


wb.save(excel_path)